"""Команда загрузит Сотрудников в модель `Employee` из xlsx файла.

| Фамилия Имя Отчество |
| Фамилия Имя Отчество |
"""

import logging
import os
from typing import Dict, List

from django.core.management.base import BaseCommand
from django.db import IntegrityError
from django.utils.text import slugify
from openpyxl import load_workbook
from transliterate import translit
from vacation_visualiser.api.employee.models import Employee


logging.basicConfig(
    filename='import_employees_log.txt',
    level=logging.INFO,
)


class Command(BaseCommand):
    """Команда работает только с файлом employees.xlsx. FIX."""

    def handle(self, *args: List[str], **options: Dict[str, str]):
        """Обработчик xlsx."""
        wb = load_workbook(
            f'{os.path.dirname(os.path.abspath(__file__))}/employees.xlsx',
        ).active

        for item in (item[0].value for item in wb.iter_rows()):
            try:
                last_name, first_name, middle_name = item.split()
                username = slugify(
                    translit(
                        f'{last_name} {first_name}', 'ru', reversed=True,
                    ),
                )
                Employee(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    middle_name=middle_name,
                    password=last_name,
                ).save()
            except AttributeError as err:
                logging.info((
                    f'Ошибка импорта: {item}, возможно пустая строка.\n'
                    f'Текст ошибки: {err} \n'
                ))
            except ValueError as err:
                logging.info((
                    f'Ошибка импорта: {item}, возможно не полное ФИО.\n'
                    f'Текст ошибки: {err} \n'
                ))

            except IntegrityError as err:
                logging.info((
                    f'Ошибка импорта: {last_name, first_name, middle_name}. \n'
                    f'Возможно сотрудник уже существует. \n'
                    f'Текст ошибки: {err} \n'
                ))
