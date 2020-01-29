"""Парсит файл лежащий в этой же директории."""

from datetime import datetime, timedelta
import logging
import os
from typing import Dict, List, Tuple

from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.cell import Cell
from vacation_visualiser.api.employee.models import Employee
from vacation_visualiser.api.vacation.models import Vacation


logging.basicConfig(
    filename='import_vacations_log.txt',
    level=logging.INFO,
)


class Command(BaseCommand):
    """Команда загрузки xlsx файла с данными об отпуске."""

    def __init__(self):
        """Открывает файл на чтение в виде генератора строк."""
        self.wb = load_workbook(
            f'{os.path.dirname(os.path.abspath(__file__))}/vacations.xlsx',
        ).active
        self.iter_rows = self.wb.iter_rows()
        super().__init__()

    def handle(self, *args: List[str], **options: Dict[str, str]) -> None:
        """Итерируется по строкам. Вызывает парсер и создание отпуска."""
        for item in (item for item in self.iter_rows):
            person_dict, vacation_dict = parse(item=item)
            if person_dict and vacation_dict:
                create_vacation(
                    person_dict=person_dict,
                    vacation_dict=vacation_dict,
                )
            else:
                logging.info('Не удалось распарсить запись')


def create_vacation(person_dict: Dict[str, str], vacation_dict: Dict[str, str]):
    """Создаёт отпуск сотрудника из корректных данных."""
    try:
        employee = Employee.objects.get(
            **person_dict,
        )
    except Employee.DoesNotExist:
        return
    employee.vacation_set.add(
        Vacation(**vacation_dict), bulk=False,
    )


def parse(item: Tuple[Cell, ...]):
    """Парсит и приводит к нужному виду определённые ячейки в строке."""
    try:
        last_name, first_name, middle_name = item[2].value.split()
        days = int(item[4].value) if item[4].value.isnumeric() else False
        date_start = datetime.strptime(item[6].value, '%d.%m.%Y').date()
        date_end = date_start + timedelta(days=days)

        return {
            'last_name': last_name,
            'first_name': first_name,
            'middle_name': middle_name,
        }, {
            'date_end': date_end,
            'date_start': date_start,
        }
    except (AttributeError, TypeError, ValueError):
        return {}, {}
